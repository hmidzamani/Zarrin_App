import snap7
import time
from openpyxl import Workbook, load_workbook

PLC_IP = "192.168.10.38"
RACK = 0
SLOT = 1
EXCEL_PATH = "plc_live_data.xlsx"

DBS = {
    "C1": 134,
    "C2": 135
}

TAGS = {
    "TotalWorktime": 8,
    "TotalProducts": 20,
    "TotalGoodProducts": 24,
    "TotalScrapProducts": 32,
    "MachineSpeed": 132
}

CELL_MAP = {
    "C1": {
        "TotalWorktime": "B2",
        "TotalProducts": "B3",
        "TotalGoodProducts": "B4",
        "TotalScrapProducts": "B5",
        "MachineSpeed": "B6",
        "Scrap_Percentage": "B7",
        "OEE": "B8"
    },
    "C2": {
        "TotalWorktime": "C2",
        "TotalProducts": "C3",
        "TotalGoodProducts": "C4",
        "TotalScrapProducts": "C5",
        "MachineSpeed": "C6",
        "Scrap_Percentage": "C7",
        "OEE": "C8"
    }
}

def read_dint(db_data, offset):
    return int.from_bytes(db_data[offset:offset+4], byteorder='big', signed=True)

def read_tags(client, db_number):
    db_data = client.db_read(db_number, 0, 136)
    return {tag: read_dint(db_data, offset) for tag, offset in TAGS.items()}

def calculate_metrics(data):
    scrap = data["TotalScrapProducts"]
    total = data["TotalProducts"]
    good = data["TotalGoodProducts"]
    worktime = data["TotalWorktime"]

    scrap_pct = round((100 * scrap / total) if total else 0, 2)
    oee = round((100 * good / (900 * worktime)) if worktime else 0, 2)
    return scrap_pct, oee

def init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "LiveData"
    for prefix in DBS:
        for tag, cell in CELL_MAP[prefix].items():
            ws[cell] = f"{prefix}_{tag}"  # Label headers
    wb.save(EXCEL_PATH)

def write_to_excel(values):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["LiveData"]
    for prefix, data in values.items():
        for tag, value in data.items():
            cell = CELL_MAP[prefix][tag]
            ws[cell] = value
    wb.save(EXCEL_PATH)

def main():
    client = snap7.client.Client()
    client.connect(PLC_IP, RACK, SLOT)
    init_excel()

    while True:
        all_values = {}
        for prefix, db_num in DBS.items():
            raw = read_tags(client, db_num)
            scrap_pct, oee = calculate_metrics(raw)
            raw["Scrap_Percentage"] = scrap_pct
            raw["OEE"] = oee
            all_values[prefix] = raw

            print(f"{prefix} →", {k: raw[k] for k in TAGS})
            print(f"{prefix} Scrap%: {scrap_pct:.2f}, OEE: {oee:.2f}")

        write_to_excel(all_values)
        time.sleep(1)

if __name__ == "__main__":
    main()
