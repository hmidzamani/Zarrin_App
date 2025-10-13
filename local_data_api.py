from flask import Flask, request, jsonify
from openpyxl import load_workbook
from flask_cors import CORS

EXCEL_PATH = "plc_live_data.xlsx"
LABELS = [
    "TotalWorktime",
    "TotalProducts",
    "TotalGoodProducts",
    "TotalScrapProducts",
    "MachineSpeed",
    "Scrap_Percentage",
    "OEE"
]

app = Flask(__name__)
CORS(app)

@app.route("/data")
def get_data():
    machine = request.args.get("machine", "C1")
    col = 1 if machine == "C1" else 2  # Column B = 1, Column C = 2

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active

        result = {}
        for row_index, label in enumerate(LABELS, start=1):
            value = ws.cell(row=row_index, column=col + 1).value  # openpyxl is 1-based
            result[label] = value

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(port=6060, debug=True)
